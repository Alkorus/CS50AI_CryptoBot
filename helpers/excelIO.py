import openpyxl
import datetime
import os

from openpyxl.workbook.workbook import Workbook

class ExcelIO:

    RECORDS_PATH = '../records'

    def __init__(self):
        print('test')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws1 = wb.create_sheet("MySheet")
        wb.save(f"{self.RECORDS_PATH}/test.xlsx")

        # wb2 = openpyxl.load_workbook('./test/test.xlsx')
        # print(wb2.sheetnames)

    @staticmethod
    def enter_record(**kwargs):
        # The one higher isn't accessible if the class isn't declared prior
        RECORDS_PATH = './records'

        # if a titlefor the report isn't given, generate one
        if 'title' in kwargs:
            title = kwargs['title']
        else:
            today = datetime.date.today()
            title = f"{today.year}-{today.month}.xlsx"

        if 'sheet' in kwargs:
            sheet = kwargs['sheet']
        else:
            today = datetime.date.today()
            sheet = f"{today.year}-{today.month}-{today.day}"

        if 'entry' in kwargs:
            entry = kwargs['entry']
        else:
            entry = {'test': 'test'}

        report_path = f"{RECORDS_PATH}/{title}"

        # check if the report file already exists
        if os.path.isfile(report_path):
            #print("report exists")
            #print(report_path)
            wb = openpyxl.load_workbook(report_path)
        else:
            #print("report doesn't exist")
            wb = openpyxl.Workbook()

        # check if sheet exists inside workbook
        if sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.create_sheet(sheet)
            # add headers for sheet
            i = 1
            for k, v in entry.items():
                d = ws.cell(row=1, column=i, value=k)
                i+=1

        # print(ws.cell(row=1, column=1).value)
        
        # register the entry
        if entry.get('test') == 'test':
            d = ws.cell(row=ws.max_row+1, column=1, value='test')
        else:
            row = ws.max_row+1
            for k, v in entry.items():
                column = ExcelIO.find_row(ws, k)
                # print(f'row: {row}, column: {column}, value: {v}')
                ws.cell(row=row, column=column, value=v)

        wb.save(report_path)

    @staticmethod
    def find_row(ws, header):
        found = False
        
        for i in range(1, ws.max_column+1): # ranges ends at value before second argument, so add 1 to include all columns
            # print(f'{i}, {ws.cell(row=1, column=i).value()}, {header}')
            if ws.cell(row=1, column=i).value == header:
                found = True
                return i
        if not found:
            i+=1
            ws.cell(row=1, column=i, value=header)
            return i






def main():
    record = {'crypto': 'BTC', 'entry time': datetime.datetime.now() + datetime.timedelta(minutes=-5), 'entry price': 40000, 'exit time': datetime.datetime.now(), 'exit price': 42000, 'relative gain': 5, 'test': 'test2'}
    line = {'title': 'test2.xlsx', 'entry':record, 'sheet': 'test2'}
    print(len(record))

    # record = {'entry':'B'}
    ExcelIO.enter_record(**line)
    # ExcelIO()


if __name__ == '__main__': main()
